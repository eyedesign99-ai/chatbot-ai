import os
import json
import sqlite3
import pickle
from datetime import datetime
from openpyxl import Workbook, load_workbook
from uuid import uuid4

from openai import OpenAI
import numpy as np

# =========================
# 1. CẤU HÌNH ĐƯỜNG DẪN
# =========================

BASE_DIR = r"F:\Chatbot\central_data"

SQLITE_PATH = os.path.join(BASE_DIR, "sqlite", "paintings.db")
TOPIC_META_PATH = os.path.join(BASE_DIR, "topics", "topic_meta.pkl")
TOPIC_VECTORS_PATH = os.path.join(BASE_DIR, "topics", "topic_vectors.npy")

LOG_DIR = r"F:\Chatbot\logs"

# =========================
# 2. OPENAI CLIENT & API KEY
# =========================

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise RuntimeError(
        "❌ Chưa thiết lập biến môi trường OPENAI_API_KEY. "
        "Hãy set trong hệ thống trước khi chạy chatbot."
    )

client = OpenAI()  # tự đọc OPENAI_API_KEY từ env

EMBED_MODEL = "text-embedding-3-small"  # 1536 chiều
CHAT_MODEL = "gpt-4o-mini"

# =========================
# 3. LOAD TOPIC INDEX
# =========================

TOPIC_VECTORS = None
TOPIC_META = None

def load_topic_index():
    global TOPIC_VECTORS, TOPIC_META
    if TOPIC_VECTORS is not None and TOPIC_META is not None:
        return

    if not (os.path.exists(TOPIC_META_PATH) and os.path.exists(TOPIC_VECTORS_PATH)):
        print("⚠ Không tìm thấy topic index, sẽ bỏ qua semantic topic search.")
        TOPIC_VECTORS = None
        TOPIC_META = []
        return

    with open(TOPIC_META_PATH, "rb") as f:
        TOPIC_META = pickle.load(f)

    TOPIC_VECTORS = np.load(TOPIC_VECTORS_PATH).astype("float32")
    print(f"✅ Đã nạp topic index: {TOPIC_VECTORS.shape[0]} topic")

# =========================
# 4. LOG CHAT VÀO EXCEL
# =========================

SESSION_ID = str(uuid4())[:8]

def log_chat(user_input, bot_reply):
    os.makedirs(LOG_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    log_file = os.path.join(LOG_DIR, f"{today}.xlsx")

    if os.path.exists(log_file):
        wb = load_workbook(log_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID phiên", "Thời gian", "Người dùng", "Chatbot"])

    current_time = datetime.now().strftime("%H:%M:%S")
    ws.append([SESSION_ID, current_time, user_input, bot_reply])
    wb.save(log_file)

# =========================
# 5. HÀM TIỆN ÍCH
# =========================

def get_db_connection():
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def embed_text(text: str):
    resp = client.embeddings.create(
        model=EMBED_MODEL,
        input=text
    )
    return np.array(resp.data[0].embedding, dtype="float32")

def normalize_query_for_like(q: str) -> str:
    return f"%{q.strip()}%"

# =========================
# 6. SEARCH SQLITE THEO KEYWORD
# =========================

def keyword_search_paintings(user_input: str, limit: int = 20):
    conn = get_db_connection()
    cur = conn.cursor()

    like_pattern = normalize_query_for_like(user_input)

    query = """
        SELECT id, title, image, general_info, keywords, themes, emotions,
               description_short, json_path
        FROM paintings
        WHERE 
            title LIKE ? OR
            keywords LIKE ? OR
            themes LIKE ? OR
            emotions LIKE ?
        LIMIT ?;
    """

    cur.execute(query, (like_pattern, like_pattern, like_pattern, like_pattern, limit))
    rows = cur.fetchall()
    conn.close()

    results = []
    for r in rows:
        results.append({
            "id": r["id"],
            "title": r["title"],
            "image": r["image"],
            "general_info": r["general_info"],
            "keywords": (r["keywords"] or "").split(",") if r["keywords"] else [],
            "themes": (r["themes"] or "").split(",") if r["themes"] else [],
            "emotions": (r["emotions"] or "").split(",") if r["emotions"] else [],
            "description_short": r["description_short"],
            "json_path": r["json_path"]
        })

    return results

# =========================
# 7. SEMANTIC TOPIC SEARCH
# =========================

def semantic_topic_search(user_input: str, top_k_topics: int = 3, max_items: int = 20):
    load_topic_index()
    if TOPIC_VECTORS is None or len(TOPIC_META) == 0:
        return []

    q_vec = embed_text(user_input)
    # chuẩn hóa
    q_norm = q_vec / (np.linalg.norm(q_vec) + 1e-8)
    topic_norm = TOPIC_VECTORS / (np.linalg.norm(TOPIC_VECTORS, axis=1, keepdims=True) + 1e-8)

    scores = topic_norm @ q_norm  # cosine similarity
    top_idx = np.argsort(scores)[::-1][:top_k_topics]

    # gom id tranh từ các topic gần nhất
    candidate_ids = []
    for idx in top_idx:
        meta = TOPIC_META[idx]
        ids = meta.get("suggest_ids", [])
        candidate_ids.extend(ids)

    # loại trùng, giữ thứ tự
    seen = set()
    unique_ids = []
    for i in candidate_ids:
        if i not in seen:
            seen.add(i)
            unique_ids.append(i)
        if len(unique_ids) >= max_items:
            break

    if not unique_ids:
        return []

    # lấy chi tiết tranh từ SQLite
    conn = get_db_connection()
    cur = conn.cursor()

    placeholders = ",".join("?" for _ in unique_ids)
    sql = f"""
        SELECT id, title, image, general_info, keywords, themes, emotions,
               description_short, json_path
        FROM paintings
        WHERE id IN ({placeholders});
    """
    cur.execute(sql, unique_ids)
    rows = cur.fetchall()
    conn.close()

    # map theo thứ tự unique_ids
    row_map = {r["id"]: r for r in rows}
    results = []
    for pid in unique_ids:
        r = row_map.get(pid)
        if not r:
            continue
        results.append({
            "id": r["id"],
            "title": r["title"],
            "image": r["image"],
            "general_info": r["general_info"],
            "keywords": (r["keywords"] or "").split(",") if r["keywords"] else [],
            "themes": (r["themes"] or "").split(",") if r["themes"] else [],
            "emotions": (r["emotions"] or "").split(",") if r["emotions"] else [],
            "description_short": r["description_short"],
            "json_path": r["json_path"]
        })

    return results

# =========================
# 8. ENRICH DATA (THÊM HTML HÌNH & LINK)
# =========================

def enrich_product_data(context_list):
    for item in context_list:
        if isinstance(item, dict) and "image" in item and "id" in item:
            img_path = item["image"]  # vd: "cgi/28.jpg"
            sp_id = item["id"]        # vd: 28

            item["hình_html"] = (
                f"<img src='/static/product/{img_path}' "
                f"style='max-width: 100%; border-radius: 10px;'>"
            )
            item["link_html"] = (
                f"<p><a href='https://cgi.vn/san-pham/{sp_id}' "
                f"target='_blank'>Xem chi tiết sản phẩm</a></p>"
            )
    return context_list

# =========================
# 9. PROMPT & GỌI GPT
# =========================

SYSTEM_PROMPT = """
Bạn là một nhân viên tư vấn bán tranh.

MỤC TIÊU HIỂN THỊ:
1) Phần trên: chỉ giới thiệu NGẮN GỌN (2–4 câu) về chủ đề tranh phù hợp với yêu cầu của khách.
2) Phần dưới: HIỂN THỊ DANH SÁCH NHIỀU TRANH (tất cả tranh có trong dữ liệu đầu vào), dạng gallery.
   - Mỗi tranh:
     - Hiển thị tiêu đề (title)
     - Hiển thị hình HTML đã cung cấp trong trường "hình_html"
     - Hiển thị link HTML đã cung cấp trong trường "link_html"
   - KHÔNG viết mô tả dài cho từng tranh. Nếu cần, chỉ 1 câu rất ngắn.

QUY TẮC QUAN TRỌNG:
- LUÔN sử dụng đúng "hình_html" và "link_html" có trong dữ liệu, KHÔNG tự bịa đường dẫn.
- KHÔNG dùng markdown kiểu ![ ](...) hoặc link giả #.
- KHÔNG tự ý rút gọn danh sách tranh xuống còn 3–5 bức.
  → Hãy hiển thị ĐẦY ĐỦ TẤT CẢ các tranh trong mảng dữ liệu đã cung cấp.
- Có thể trình bày danh sách tranh theo dạng:

  <h3>Danh sách tranh gợi ý</h3>
  <div class="gallery">
    <!-- lặp qua từng tranh -->
    <div class="item">
      <h4>{title}</h4>
      {hình_html}
      {link_html}
    </div>
  </div>

- Giữ giọng văn thân thiện, dễ hiểu, nhưng ưu tiên NGẮN GỌN.
"""

def query_openai_with_context(context_list, user_input):
    context_list = enrich_product_data(context_list)
    context_text = json.dumps(context_list, ensure_ascii=False, indent=2)

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"Khách hỏi: {user_input}\n\nDữ liệu sản phẩm:\n{context_text}"}
    ]

    resp = client.chat.completions.create(
        model=CHAT_MODEL,
        messages=messages,
        temperature=0.7
    )

    return resp.choices[0].message.content

# =========================
# 10. ROUTER SEARCH: KEYWORD + SEMANTIC
# =========================

def search_paintings_for_user_query(user_input: str, max_results: int = 20):
    # B1: thử keyword search trước
    kw_results = keyword_search_paintings(user_input, limit=max_results)
    if len(kw_results) >= 8:
        print(f"🔎 Keyword search trả {len(kw_results)} kết quả → dùng trực tiếp.")
        return kw_results

    # B2: nếu keyword ít kết quả → dùng semantic topic search
    print(f"🔍 Keyword chỉ có {len(kw_results)} kết quả → dùng thêm semantic topic search.")
    sem_results = semantic_topic_search(user_input, top_k_topics=3, max_items=max_results)

    # Nếu semantic có kết quả → ưu tiên
    if sem_results:
        print(f"🧠 Semantic topic search trả {len(sem_results)} kết quả.")
        return sem_results

    # Nếu semantic cũng không có → fallback về keyword
    print("⚠ Semantic topic search không có kết quả → fallback keyword.")
    return kw_results

# =========================
# 11. MAIN CHATBOT LOOP
# =========================

def chatbot():
    print("🤖 Chatbot đã sẵn sàng! Gõ 'exit' để thoát.\n")
    while True:
        user_input = input("Bạn: ")
        if user_input.lower().strip() == "exit":
            print("👋 Chatbot kết thúc.")
            break

        try:
            context_list = search_paintings_for_user_query(user_input)
            gpt_reply = query_openai_with_context(context_list, user_input)
        except Exception as e:
            print("❌ Lỗi trong quá trình xử lý:", e)
            gpt_reply = "Xin lỗi, hiện tại mình đang gặp chút trục trặc hệ thống, bạn thử lại sau nhé."

        print("Chatbot:", gpt_reply)
        log_chat(user_input, gpt_reply)

if __name__ == "__main__":
    chatbot()
